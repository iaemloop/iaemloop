#!/usr/bin/env python3
"""
Extrai ratings de bancos de um arquivo Excel (baixado manualmente da XP).
Estrutura conhecida: colunas 1 (emissor), 2 (Fitch rating), 4 (S&P rating), 6 (Moody's rating).
"""

import json
import os
import sys

EXCEL_PATH = 'data/ratings.xlsx'
OUTPUT_PATH = 'data/ratings_bancos.json'

def extract_ratings():
    try:
        import openpyxl
    except ImportError:
        print("❌ Instale openpyxl: python3 -m pip install openpyxl --break-system-packages")
        return None

    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Arquivo não encontrado: {EXCEL_PATH}")
        return None

    print(f"📖 Lendo {EXCEL_PATH}...")
    workbook = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet = workbook.active

    # Encontrar linha do cabeçalho (contém 'Razão Social do Emissor')
    header_row_idx = None
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if any(cell and 'Razão Social' in str(cell) for cell in row):
            header_row_idx = row_idx
            break

    if header_row_idx is None:
        print("❌ Cabeçalho não encontrado")
        return None

    print(f"   Cabeçalho na linha {header_row_idx}")

    # Estrutura fixa baseada na análise:
    # col 1: Razão Social do Emissor
    # col 2: FITCH RATINGS BRASIL LTDA. (rating)
    # col 3: Perspectiva Rating Fitch
    # col 4: STANDARD & POOR'S RATINGS DO BRASIL LTDA. (rating)
    # col 5: Perspectiva Rating S&P
    # col 6: MOODY'S LOCAL BRASIL (rating)
    # col 7: Perspectiva Rating Moody's
    col_emissor = 1
    col_fitch = 2
    col_sp = 4
    col_moody = 6

    print(f"   Usando colunas fixas: Emissor={col_emissor}, Fitch={col_fitch}, S&P={col_sp}, Moody={col_moody}")

    # Extrair dados
    ratings = {}
    row_count = 0
    for row in sheet.iter_rows(min_row=header_row_idx+1, values_only=True):
        row_count += 1
        if not row or not any(row):
            continue

        emissor = str(row[col_emissor]).strip() if len(row) > col_emissor else ''
        if not emissor or emissor.lower() in ('n/a', '-', 'nan', ''):
            continue

        def clean(val):
            if val is None:
                return None
            s = str(val).strip()
            if s.lower() in ('n/a', '-', 'nan', ''):
                return None
            return s

        fitch = clean(row[col_fitch]) if len(row) > col_fitch else None
        sp = clean(row[col_sp]) if len(row) > col_sp else None
        moody = clean(row[col_moody]) if len(row) > col_moody else None

        ratings[emissor] = {
            "moody": moody,
            "fitch": fitch,
            "sp": sp
        }

    print(f"✅ {len(ratings)} instituições extraídas de {row_count} linhas")
    return ratings

def main():
    ratings = extract_ratings()
    if not ratings:
        sys.exit(1)

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(ratings, f, ensure_ascii=False, indent=2)

    print(f"📄 Ratings salvos em {OUTPUT_PATH}")
    print("💡 Execute update_fgc_ranking.py para aplicar ao ranking.")

if __name__ == '__main__':
    main()
