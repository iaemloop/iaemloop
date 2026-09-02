#!/usr/bin/env python3
"""
Baixa e extrai ratings de bancos do Excel da XP.
Fonte: https://conteudos.xpi.com.br/wp-content/uploads/2026/02/RatingsExternosEmissores-XPResearchRF-20260130-1.xlsx
"""

import requests
import json
import os
import sys
from datetime import datetime

# URL do Excel (sorted by date: 20260130 = Jan 30, 2026)
EXCEL_URL = "https://conteudos.xpi.com.br/wp-content/uploads/2026/02/RatingsExternosEmissores-XPResearchRF-20260130.xlsx"

OUTPUT_PATH = 'data/ratings_bancos.json'
TEMP_EXCEL = 'data/ratings_temp.xlsx'

def download_excel():
    print(f"🔄 Baixando Excel de {EXCEL_URL}...")
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream,*/*",
        "Referer": "https://conteudos.xpi.com.br/renda-fixa/relatorios/guia-de-rating/",
    }
    try:
        response = requests.get(EXCEL_URL, headers=headers, timeout=60)
        response.raise_for_status()
        os.makedirs('data', exist_ok=True)
        with open(TEMP_EXCEL, 'wb') as f:
            f.write(response.content)
        print(f"✅ Download completo: {len(response.content)} bytes")
        return True
    except Exception as e:
        print(f"❌ Erro ao baixar: {e}")
        return False

def parse_excel():
    try:
        import openpyxl
    except ImportError:
        print("❌ Módulo openpyxl não encontrado. Instale com: pip install openpyxl")
        return None

    print("📖 Lendo Excel...")
    try:
        workbook = openpyxl.load_workbook(TEMP_EXCEL, data_only=True)
        sheet = workbook.active
    except Exception as e:
        print(f"❌ Erro ao ler Excel: {e}")
        return None

    # Descobrir estrutura: procurar cabeçalhos
    headers = []
    for row in sheet.iter_rows(max_row=1, values_only=True):
        headers = [str(h).strip() if h else '' for h in row]
        break

    print(f"   Cabeçalhos encontrados: {headers}")

    # Mapear colunas relevantes
    # Procurar por: 'Banco', 'Issuer', 'Emissor', 'Moody', 'Fitch', 'S&P', 'SP', 'Rating'
    col_emissor = None
    col_moody = None
    col_fitch = None
    col_sp = None

    for i, h in enumerate(headers):
        h_lower = h.lower()
        if 'banco' in h_lower or 'issuer' in h_lower or 'emissor' in h_lower:
            col_emissor = i
        if 'moody' in h_lower:
            col_moody = i
        if 'fitch' in h_lower:
            col_fitch = i
        if 's&p' in h_lower or 'sp' in h_lower or 'stand' in h_lower:
            col_sp = i

    if col_emissor is None:
        print("❌ Coluna do emissor/banco não encontrada")
        return None

    print(f"   Colunas mapeadas: Emissor={col_emissor}, Moody={col_moody}, Fitch={col_fitch}, S&P={col_sp}")

    # Extrair linhas
    ratings = {}
    row_count = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_count += 1
        if not row or not any(row):
            continue

        emissor = str(row[col_emissor]).strip() if col_emissor is not None and len(row) > col_emissor else ''
        if not emissor or emissor.lower() in ('n/a', '-', 'nan'):
            continue

        # Pegar ratings (prioridade: Fitch > Moody's > S&P)
        moody = str(row[col_moody]).strip() if col_moody is not None and len(row) > col_moody and row[col_moody] else None
        fitch = str(row[col_fitch]).strip() if col_fitch is not None and len(row) > col_fitch and row[col_fitch] else None
        sp = str(row[col_sp]).strip() if col_sp is not None and len(row) > col_sp and row[col_sp] else None

        # Limpar valores
        def clean(val):
            if val is None or str(val).lower() in ('n/a', '-', 'nan', ''):
                return None
            return str(val).strip()

        moody = clean(moody)
        fitch = clean(fitch)
        sp = clean(sp)

        ratings[emissor] = {
            "moody": moody,
            "fitch": fitch,
            "sp": sp
        }

    print(f"✅ {len(ratings)} instituições extraídas de {row_count} linhas")
    return ratings

def main():
    if not download_excel():
        sys.exit(1)

    ratings = parse_excel()
    if not ratings:
        print("❌ Falha ao extrair dados do Excel")
        sys.exit(1)

    # Salvar JSON
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(ratings, f, ensure_ascii=False, indent=2)

    print(f"📄 Ratings salvos em {OUTPUT_PATH}")
    print("💡 Agora execute update_fgc_ranking.py para aplicar ao ranking.")

    # Limpar temporário
    try:
        os.remove(TEMP_EXCEL)
    except:
        pass

if __name__ == '__main__':
    main()
