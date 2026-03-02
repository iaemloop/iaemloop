#!/usr/bin/env python3
"""
Extrai ratings do Excel da XP (formato fixo Columns A,H,...).
Estrutura observada:
- Linha 4: cabeçalhos (col 1 = Razão Social do Emissor, col 2 = nome agência Fitch, col 3 = Perspectiva Fitch, col 4 = nome agência S&P, col 5 = Perspectiva S&P, col 6 = nome agência Moody's, col 7 = Perspectiva Moody's)
- Dados a partir da linha 5.
- Na verdade os ratings aparecem nas colunas 2, 4, 6 (mesmas colunas dos nomes das agências).
- Perspectivas nas colunas 3, 5, 7.
"""

import json
import os
import sys

try:
    import openpyxl
except ImportError:
    print("❌ Instale openpyxl: python3 -m pip install openpyxl --break-system-packages")
    sys.exit(1)

EXCEL_PATH = 'data/ratings.xlsx'
OUTPUT_PATH = 'data/ratings_bancos.json'

if not os.path.exists(EXCEL_PATH):
    print(f"❌ Arquivo não encontrado: {EXCEL_PATH}")
    sys.exit(1)

print(f"📖 Lendo {EXCEL_PATH}...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb.active

# Achar linha do cabeçalho (onde está 'Razão Social do Emissor')
header_row = None
for i in range(1, 50):
    row = next(ws.iter_rows(min_row=i, max_row=i, values_only=True))
    if any(cell and 'Razão Social' in str(cell) for cell in row):
        header_row = i
        break

if header_row is None:
    print("❌ Cabeçalho não encontrado")
    sys.exit(1)

print(f"   Cabeçalho na linha {header_row}")

# Índices fixos (0-based)
IDX_EMISSOR = 1   # col 2 (Razão Social)
IDX_FITCH = 2     # col 3 (FITCH RATINGS... -> rating)
IDX_SP = 4        # col 5 (S&P -> rating)
IDX_MOODY = 6     # col 7 (MOODY'S -> rating)

# Extrair a partir da próxima linha
ratings = {}
count = 0
for row in ws.iter_rows(min_row=header_row+1, values_only=True):
    if not row or not any(row):
        continue
    emissor = str(row[IDX_EMISSOR]).strip() if len(row) > IDX_EMISSOR and row[IDX_EMISSOR] else None
    if not emissor:
        continue

    def clean(val):
        if val is None:
            return None
        s = str(val).strip()
        if s.lower() in ('n/a', '-', 'nan', ''):
            return None
        return s

    fitch = clean(row[IDX_FITCH]) if len(row) > IDX_FITCH else None
    sp = clean(row[IDX_SP]) if len(row) > IDX_SP else None
    moody = clean(row[IDX_MOODY]) if len(row) > IDX_MOODY else None

    ratings[emissor] = {
        "moody": moody,
        "fitch": fitch,
        "sp": sp
    }
    count += 1

print(f"✅ {len(ratings)} instituições extraídas")

# Salvar JSON
os.makedirs(os.path.dirname(OUTPUT_PATH) or '.', exist_ok=True)
with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    json.dump(ratings, f, ensure_ascii=False, indent=2)

print(f"📄 Ratings salvos em {OUTPUT_PATH}")
print("💡 Agora execute: python3 scripts/update_fgc_ranking.py")
